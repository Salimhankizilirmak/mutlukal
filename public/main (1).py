from openpyxl import load_workbook, Workbook

# --- CHECK DIGIT (GS1 MOD10) ---
def calculate_check_digit(number):
    digits = [int(d) for d in number]
    total = 0
    reverse = digits[::-1]

    for i, d in enumerate(reverse):
        if i % 2 == 0:
            total += d * 3
        else:
            total += d

    return (10 - (total % 10)) % 10


# --- AYARLAR ---
INPUT_FILE = "rapor.xlsx"
OUTPUT_FILE = "sonuc2.xlsx"

GLN = "869882938"
EXTENSION = "2"

ITEMS_PER_CARTON = 30
CARTONS_PER_PALLET = 84

serial = 1000000


# --- INPUT OKU ---
wb_in = load_workbook(INPUT_FILE)
ws_in = wb_in.active

products = []
for row in ws_in.iter_rows(values_only=True):
    if row and row[0]:
        products.append(str(row[0]))


# --- OUTPUT ---
wb_out = Workbook()
ws_out = wb_out.active

ws_out.append(["UrunBarkod", "KoliSSCC", "PaletSSCC"])


current_carton_sscc = ""
current_pallet_sscc = ""

for i, product in enumerate(products, start=1):

    # YENİ KOLİ
    if (i - 1) % ITEMS_PER_CARTON == 0:
        serial_str = str(serial).zfill(7)
        base = EXTENSION + GLN + serial_str
        check = calculate_check_digit(base)

        current_carton_sscc = f"(00){base}{check}"
        serial += 1

    # YENİ PALET
    carton_no = (i - 1) // ITEMS_PER_CARTON + 1
    if (carton_no - 1) % CARTONS_PER_PALLET == 0 and (i - 1) % ITEMS_PER_CARTON == 0:
        serial_str = str(serial).zfill(7)
        base = EXTENSION + GLN + serial_str
        check = calculate_check_digit(base)

        current_pallet_sscc = f"(00){base}{check}"
        serial += 1

    # HER SATIRA YAZ
    ws_out.append([
        product,
        current_carton_sscc,
        current_pallet_sscc
    ])


# --- KAYDET ---
wb_out.save(OUTPUT_FILE)

print("Tam dolu Excel oluşturuldu:", OUTPUT_FILE)